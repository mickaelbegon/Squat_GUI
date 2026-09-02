"""Versioned, shared export schema for CSV and Excel outputs."""

from __future__ import annotations

from collections import OrderedDict
from copy import copy
from dataclasses import dataclass
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
from typing import Iterable, Mapping, Sequence

SCHEMA_VERSION = "2.0.0"
SEGMENTS = ("foot", "shank", "thigh", "trunk", "bar")
JOINTS = ("cheville", "genou", "hanche")
POINTS = ("heel", "toe", "ankle", "knee", "hip", "shoulder", "bar")
ROW_KEYS = ("schema_version", "condition_id", "frame", "time_s")
SUMMARY_SHEET = "Synthèse"
COMBINED_SHEET = "Données combinées"
DEFINITIONS_SHEET = "Définitions"


@dataclass(frozen=True)
class ColumnDefinition:
    unit: str
    definition: str
    sign_convention: str = "sans objet"
    status: str = "canonique"


CONDITION_COLUMNS = (
    "schema_version",
    "condition_id",
    "backend",
    "subject_profile",
    "bar_position",
    "wedge_20_deg",
    "load_percent_bw",
    "load_kg",
    "body_mass_kg",
    "total_mass_kg",
    "height_m",
    "shank_percent",
    "thigh_percent",
    "trunk_percent",
    "anthropometry_mode",
    "anthropometry_scaling_rule",
    "duration_excentrique_s",
    "duration_isometrique_s",
    "duration_concentrique_s",
    "total_duration_s",
    "frames",
    "torque_preset",
    "angle_adapt",
    "velocity_adapt",
    "max_cheville_Nm",
    "max_genou_Nm",
    "max_hanche_Nm",
)

TIME_COLUMNS = ROW_KEYS + (
    "delta_time_s",
    "normalized_time_percent",
    "phase",
    "backend",
)
COORDINATE_COLUMNS = ROW_KEYS + tuple(
    f"{point}_{axis}_m" for point in POINTS for axis in ("x", "y")
)
ORIENTATION_COLUMNS = ROW_KEYS + tuple(
    f"{segment}_orientation_deg" for segment in ("foot", "shank", "thigh", "trunk")
)
KINEMATIC_COLUMNS = (
    ROW_KEYS
    + (
        "q_shank_deg",
        "q_thigh_deg",
        "q_trunk_deg",
    )
    + tuple(
        f"{joint}_{quantity}{unit}"
        for joint in JOINTS
        for quantity, unit in (
            ("angle", "_deg"),
            ("velocity", "_deg_s"),
            ("acceleration", "_deg_s2"),
        )
    )
)
ANTHROPOMETRY_COLUMNS = ("schema_version", "condition_id", "segment") + (
    "mass_kg",
    "mass_fraction_body",
    "length_m",
    "com_fraction",
    "com_transverse_offset_m",
    "radius_of_gyration_fraction",
    "inertia_kg_m2",
    "scaling_mode",
    "scaling_rule",
    "attachment_anterior_offset_m",
    "attachment_longitudinal_offset_m",
)
SEGMENT_COM_COLUMNS = ROW_KEYS + tuple(
    f"{segment}_{quantity}"
    for segment in SEGMENTS
    for quantity in (
        "com_x_m",
        "com_y_m",
        "weighted_com_x_kg_m",
        "weighted_com_y_kg_m",
    )
)
GLOBAL_COM_COLUMNS = ROW_KEYS + (
    "total_mass_kg",
    "com_x_m",
    "com_y_m",
    "com_vx_m_s",
    "com_vy_m_s",
    "com_ax_m_s2",
    "com_ay_m_s2",
)
FORCE_COLUMNS = ROW_KEYS + (
    "grf_x_N",
    "grf_y_N",
    "weight_magnitude_N",
    "weight_x_N",
    "weight_y_N",
    "force_balance_residual_x_N",
    "force_balance_residual_y_N",
    "support_point_x_m",
    "support_point_label",
    "support_point_source",
    "geometric_support_posterior_m",
    "geometric_support_anterior_m",
    "functional_support_posterior_m",
    "functional_support_anterior_m",
    "support_point_geometric_posterior_margin_m",
    "support_point_geometric_anterior_margin_m",
    "support_point_functional_posterior_margin_m",
    "support_point_functional_anterior_margin_m",
    "support_point_in_geometric_base",
    "support_point_in_functional_base",
    "com_projection_geometric_posterior_margin_m",
    "com_projection_geometric_anterior_margin_m",
    "com_projection_in_geometric_base",
)
DYNAMIC_COLUMNS = (
    ROW_KEYS
    + ("dynamic_moment_z_Nm", "contact_source")
    + tuple(
        f"{joint}_{quantity}"
        for joint in JOINTS
        for quantity in (
            "torque_Nm",
            "torque_body_mass_normalized_Nm_kg",
            "max_available_Nm",
            "capacity_base_torque_Nm",
            "capacity_angle_rad",
            "capacity_angular_velocity_rad_s",
            "capacity_angle_factor",
            "capacity_velocity_factor",
            "capacity_regime",
            "capacity_regime_source",
            "capacity_angle_in_domain",
            "capacity_model",
            "capacity_source",
            "capacity_defined",
            "utilization_ratio",
            "utilization_percent",
            "utilization_exceeds_capacity",
            "effort_percent",
            "power_W",
            "inverse_dynamics_total_Nm",
            "mass_acceleration_Nm",
            "velocity_dependent_Nm",
            "gravity_Nm",
            "external_contact_effect_Nm",
            "inverse_dynamics_reconstruction_residual_Nm",
            "contact_Nm",
            "inertial_nonlinear_Nm",
        )
    )
)

# Stable, student-facing CSV contract.  The complete in-memory row remains the
# source for the diagnostic Excel sheets and for the opt-in ``full`` CSV mode.
STANDARD_CSV_COLUMNS = (
    "schema_version",
    "condition_id",
    "subject_profile",
    "bar_position",
    "load_percent_bw",
    "wedge_20_deg",
    "shank_percent",
    "thigh_percent",
    "trunk_percent",
    "duration_excentrique_s",
    "duration_isometrique_s",
    "duration_concentrique_s",
    "frames",
    "backend",
    "torque_preset",
    "angle_adapt",
    "velocity_adapt",
    "anthropometry_mode",
    "frame",
    "time_s",
    "delta_time_s",
    "normalized_time_percent",
    "phase",
    "q_shank_deg",
    "q_thigh_deg",
    "q_trunk_deg",
) + tuple(
    f"{joint}_{quantity}"
    for joint in JOINTS
    for quantity in (
        "angle_deg",
        "velocity_deg_s",
        "acceleration_deg_s2",
        "torque_Nm",
        "torque_body_mass_normalized_Nm_kg",
        "inverse_dynamics_total_Nm",
        "external_contact_effect_Nm",
        "inertial_nonlinear_Nm",
        "power_W",
        "utilization_ratio",
        "utilization_percent",
    )
) + (
    "com_x_m",
    "com_y_m",
    "support_point_x_m",
    "support_point_label",
    "support_point_source",
    "functional_support_posterior_m",
    "functional_support_anterior_m",
    "support_point_in_geometric_base",
    "support_point_in_functional_base",
    "grf_y_N",
)

SUMMARY_COLUMNS = (
    "schema_version",
    "condition_id",
    "subject_profile",
    "bar_position",
    "load_percent_bw",
    "wedge_20_deg",
    "shank_percent",
    "thigh_percent",
    "trunk_percent",
    "duration_excentrique_s",
    "duration_isometrique_s",
    "duration_concentrique_s",
    "frames",
    "backend",
    "torque_preset",
    "angle_adapt",
    "velocity_adapt",
    "anthropometry_mode",
    "squat_com_x_m",
    "squat_cop_x_m",
    "support_point_label",
    "zmp_x_min_m",
    "zmp_x_max_m",
    "zmp_excursion_m",
    "zmp_outside_support_frames",
    "zmp_outside_support_percent",
    "cop_outside_foot_frames",
    "cop_outside_foot_percent",
    "over_limit_frames",
    "peak_grf_y_N",
) + tuple(
    f"{joint}_{quantity}"
    for joint in JOINTS
    for quantity in (
        "peak_abs_torque_Nm",
        "peak_abs_torque_body_mass_normalized_Nm_kg",
        "peak_abs_power_W",
        "peak_utilization_ratio",
        "peak_utilization_percent",
    )
) + (
    "maximum_utilization_ratio",
    "maximum_utilization_percent",
    "limiting_joint",
    "limiting_frame",
    "limiting_time_s",
    "limiting_phase",
    "exceeds_capacity",
    "undefined_capacity_events",
)

DESCRIPTION_OVERRIDES = {
    "schema_version": "Version du contrat d'export Squat GUI.",
    "condition_id": "Identifiant stable de la condition simulée.",
    "frame": "Indice entier de l'échantillon, à partir de zéro.",
    "time_s": "Temps physique écoulé depuis le début de la simulation.",
    "delta_time_s": "Pas de temps local entre échantillons adjacents.",
    "normalized_time_percent": "Temps normalisé sur la durée totale du mouvement.",
    "phase": "Phase du mouvement: excentrique, isométrique ou concentrique.",
    "backend": "Backend ayant effectivement produit les résultats.",
    "frames": "Nombre total d'échantillons de la condition.",
    "support_point_label": "Nature du point d'appui exporté: CoP ou ZMP.",
    "support_point_source": "Méthode exacte utilisée pour calculer le point d'appui.",
    "contact_source": "Méthode effectivement utilisée pour calculer le diagnostic de contact externe.",
    "support_point_x_m": "Abscisse du CoP ou ZMP sur le plan du sol.",
    "torque_body_mass_normalized_Nm_kg": (
        "Moment articulaire divisé par la masse corporelle du sujet."
    ),
    "squat_com_x_m": (
        "Abscisse moyenne du CoM pendant la phase isométrique; à défaut, "
        "valeur à la hauteur minimale du CoM."
    ),
    "squat_cop_x_m": (
        "Abscisse moyenne du point d'appui CoP/ZMP pendant la phase "
        "isométrique; à défaut, valeur à la hauteur minimale du CoM."
    ),
    "zmp_x_min_m": "Abscisse minimale du point d'appui CoP/ZMP sur la trajectoire.",
    "zmp_x_max_m": "Abscisse maximale du point d'appui CoP/ZMP sur la trajectoire.",
    "zmp_excursion_m": "Étendue max-min du point d'appui CoP/ZMP sur la trajectoire.",
    "zmp_outside_support_frames": (
        "Nombre de frames où le point d'appui sort de la base fonctionnelle."
    ),
    "zmp_outside_support_percent": (
        "Pourcentage de frames où le point d'appui sort de la base fonctionnelle."
    ),
    "cop_outside_foot_frames": (
        "Nombre de frames où le point d'appui sort de la base géométrique du pied."
    ),
    "cop_outside_foot_percent": (
        "Pourcentage de frames où le point d'appui sort de la base géométrique du pied."
    ),
    "over_limit_frames": (
        "Nombre de frames où au moins une demande articulaire dépasse la capacité active."
    ),
    "peak_grf_y_N": "Valeur absolue maximale de la force de réaction verticale.",
    "peak_abs_torque_Nm": "Valeur absolue maximale du moment articulaire.",
    "peak_abs_torque_body_mass_normalized_Nm_kg": (
        "Valeur absolue maximale du moment articulaire normalisé par la masse corporelle."
    ),
    "peak_abs_power_W": "Valeur absolue maximale de la puissance articulaire.",
    "peak_utilization_ratio": "Valeur maximale du ratio demande/capacité U.",
    "peak_utilization_percent": "Valeur maximale de U exprimée en pourcentage.",
    "maximum_utilization_ratio": "Maximum de U parmi toutes les articulations et frames.",
    "maximum_utilization_percent": "Maximum de U exprimé en pourcentage.",
    "limiting_joint": "Articulation associée au maximum de U.",
    "limiting_frame": "Frame associée au maximum de U.",
    "limiting_time_s": "Temps associé au maximum de U.",
    "limiting_phase": "Phase associée au maximum de U.",
    "exceeds_capacity": "Vrai si une demande articulaire dépasse la capacité active.",
    "undefined_capacity_events": (
        "Nombre de demandes non nulles pour lesquelles la capacité active est nulle ou indéfinie."
    ),
    "weight_magnitude_N": "Norme du poids total, calculée comme masse totale multipliée par g.",
    "dynamic_moment_z_Nm": "Dérivée du moment cinétique autour de l'axe z global.",
    "max_available_Nm": (
        "Capacité de couple actif selon les facteurs angle et vitesse sélectionnés; "
        "elle n'inclut pas le couple passif."
    ),
    "capacity_base_torque_Nm": (
        "Amplitude de couple de base avant facteurs angle-vitesse; sa provenance "
        "est donnée par torque_preset et les colonnes max_*_Nm de la condition."
    ),
    "capacity_angle_rad": (
        "Angle fourni à Anderson: dorsiflexion/flexion positive; la flexion du "
        "genou Squat_GUI est donc inversée."
    ),
    "capacity_angular_velocity_rad_s": (
        "Vitesse dans la direction d'action du groupe testé; positive concentrique, "
        "négative excentrique."
    ),
    "capacity_angle_factor": "Multiplicateur couple-angle actif d'Anderson.",
    "capacity_velocity_factor": "Multiplicateur couple-vitesse actif d'Anderson.",
    "capacity_regime": "Régime déduit de la vitesse de capacité: concentrique, excentrique ou isométrique.",
    "capacity_regime_source": "Règle utilisée pour relier vitesse, couple, puissance et régime de capacité.",
    "capacity_angle_in_domain": "Vrai si l'angle appartient au lobe positif de la relation active d'Anderson.",
    "capacity_model": "Nom du modèle de capacité effectivement appliqué.",
    "capacity_source": "Référence primaire des paramètres de capacité.",
    "capacity_defined": "Vrai si la capacité active est strictement positive et U calculable.",
    "utilization_ratio": "U = valeur absolue du couple requis divisée par la capacité active disponible.",
    "utilization_percent": "Utilisation demande/capacité U exprimée en pourcentage.",
    "utilization_exceeds_capacity": (
        "Vrai si U > 1, ou si un couple non nul est requis alors que la capacité active est nulle."
    ),
    "effort_percent": "Alias de compatibilité de utilization_percent.",
    "inverse_dynamics_total_Nm": (
        "Couple de dynamique inverse du modèle à pied fixé, reconstruit par "
        "M(q)qddot + termes dépendant de qdot + gravité."
    ),
    "mass_acceleration_Nm": "Terme M(q)qddot projeté dans la convention des moments articulaires.",
    "velocity_dependent_Nm": "Termes de dynamique inverse dépendant de qdot, isolés à qddot nul.",
    "gravity_Nm": "Terme gravitaire de dynamique inverse, évalué à qdot=qddot=0.",
    "external_contact_effect_Nm": (
        "Effet signé du moment de GRF: opposé de la colonne legacy contact_Nm; "
        "hors reconstruction du modèle contraint à pied fixé."
    ),
    "inverse_dynamics_reconstruction_residual_Nm": (
        "Résidu total - [M(q)qddot + vitesse + gravité]; attendu nul à la tolérance numérique."
    ),
    "segment": "Nom canonique du segment ou de la barre.",
    "mass_kg": "Masse effectivement utilisée pour le segment.",
    "mass_fraction_body": "Fraction de la masse corporelle effectivement attribuée au segment; sans objet pour la barre.",
    "length_m": "Longueur effectivement utilisée pour le segment.",
    "com_fraction": "Position longitudinale du CoM en fraction de longueur segmentaire.",
    "com_transverse_offset_m": "Décalage transverse du CoM dans le repère segmentaire.",
    "radius_of_gyration_fraction": "Rayon de giration en fraction de longueur segmentaire.",
    "inertia_kg_m2": "Moment d'inertie planaire effectivement utilisé.",
    "anthropometry_mode": "Mode de sensibilité anthropométrique sélectionné pour la condition.",
    "anthropometry_scaling_rule": "Règle exacte reliant variations de longueur, masses et inerties.",
    "scaling_mode": "Mode anthropométrique appliqué à cette ligne segmentaire.",
    "scaling_rule": "Règle de recalibrage effectivement appliquée à cette ligne segmentaire.",
    "attachment_anterior_offset_m": "Décalage antérieur de l'attache de barre relativement à l'épaule.",
    "attachment_longitudinal_offset_m": "Décalage longitudinal de l'attache de barre relativement à l'épaule.",
}

LEGACY_COLUMNS = {
    "cop_x_m",
    "zmp_posterior_limit_m",
    "zmp_anterior_limit_m",
    "zmp_in_support",
    "cop_in_foot",
    "contact_Nm",
    "inertial_nonlinear_Nm",
    "effort_percent",
}


def _unit(column: str) -> str:
    suffixes = (
        ("_kg_m2", "kg·m²"),
        ("_kg_m", "kg·m"),
        ("_Nm_kg", "N·m/kg"),
        ("_deg_s2", "deg/s²"),
        ("_rad_s", "rad/s"),
        ("_rad", "rad"),
        ("_m_s2", "m/s²"),
        ("_deg_s", "deg/s"),
        ("_m_s", "m/s"),
        ("_Nm", "N·m"),
        ("_N", "N"),
        ("_kg", "kg"),
        ("_deg", "deg"),
        ("_m", "m"),
        ("_W", "W"),
        ("_percent", "%"),
        ("_fraction", "1"),
    )
    for suffix, unit in suffixes:
        if column.endswith(suffix):
            return unit
    return "1"


def _sign_convention(column: str) -> str:
    if "orientation_deg" in column:
        return "positif antihoraire depuis l'axe global +x"
    if any(
        token in column for token in ("_x_", "_vx_", "_ax_", "anterior", "posterior")
    ):
        return "+x vers l'avant; une marge positive indique l'intérieur de la limite"
    if any(token in column for token in ("_y_", "_vy_", "_ay_")):
        return "+y vers le haut"
    if any(
        token in column
        for token in (
            "torque",
            "moment",
            "contact",
            "mass_acceleration",
            "velocity_dependent",
            "gravity_Nm",
            "inverse_dynamics",
        )
    ):
        return "positif selon la convention de dynamique inverse documentée; axe +z hors du plan"
    if "power" in column:
        return "positive pour une puissance articulaire génératrice, négative pour absorbante"
    if "capacity_angular_velocity" in column:
        return "positive concentrique, négative excentrique pour le groupe musculaire modélisé"
    if any(
        token in column for token in ("angle_deg", "velocity_deg", "acceleration_deg")
    ):
        return "angles articulaires relatifs selon la convention cinématique documentée"
    return "sans objet"


def column_definition(column: str) -> ColumnDefinition:
    description = DESCRIPTION_OVERRIDES.get(column)
    if description is None:
        for joint in JOINTS:
            prefix = f"{joint}_"
            if column.startswith(prefix):
                description = DESCRIPTION_OVERRIDES.get(column[len(prefix) :])
                if description is not None:
                    description = f"{joint.capitalize()}: {description}"
                break
    if description is None:
        description = (
            column.replace("_", " ")
            .replace("com", "CoM")
            .replace("grf", "GRF")
            .capitalize()
            + "."
        )
    return ColumnDefinition(
        unit=_unit(column),
        definition=description,
        sign_convention=_sign_convention(column),
        status=(
            "compatibilité legacy"
            if column in LEGACY_COLUMNS
            or any(column.endswith(f"_{legacy}") for legacy in LEGACY_COLUMNS)
            else "canonique"
        ),
    )


def add_schema_version(rows: Iterable[Mapping[str, object]]) -> list[dict[str, object]]:
    versioned = []
    for source in rows:
        row = {"schema_version": SCHEMA_VERSION}
        row.update(source)
        versioned.append(row)
    return versioned


def csv_export_rows(
    rows: Sequence[Mapping[str, object]], *, mode: str = "standard"
) -> list[dict[str, object]]:
    """Return rows following the stable CSV contract selected by ``mode``.

    ``standard`` is the concise student-facing contract. ``full`` preserves the
    complete diagnostic row, including legacy compatibility columns.
    """
    if mode not in {"standard", "full"}:
        raise ValueError("Le mode CSV doit valoir standard ou full.")
    versioned = add_schema_version(rows)
    if mode == "full":
        return versioned
    return [
        {column: row.get(column) for column in STANDARD_CSV_COLUMNS}
        for row in versioned
    ]


def _number(row: Mapping[str, object], column: str) -> float:
    value = row.get(column)
    if value is None:
        raise ValueError(f"Valeur numérique absente: {column}")
    return float(value)


def _mean(rows: Sequence[Mapping[str, object]], column: str) -> float:
    return sum(_number(row, column) for row in rows) / len(rows)


def _frame_restarts(previous: object, current: object) -> bool:
    """Return whether two adjacent frames reveal a repeated simulation ID."""
    if previous is None or current is None:
        return False
    try:
        return float(current) <= float(previous)
    except (TypeError, ValueError):
        return False


def _simulation_groups(
    rows: Sequence[Mapping[str, object]],
) -> list[tuple[object, list[Mapping[str, object]]]]:
    """Keep contiguous simulations distinct, even when their IDs are repeated."""
    groups: list[tuple[object, list[Mapping[str, object]]]] = []
    current_id: object = None
    current_rows: list[Mapping[str, object]] = []
    previous_frame: object = None
    for row in rows:
        condition_id = row.get("condition_id")
        frame = row.get("frame")
        starts_new = bool(current_rows) and (
            condition_id != current_id or _frame_restarts(previous_frame, frame)
        )
        if starts_new:
            groups.append((current_id, current_rows))
            current_rows = []
        if not current_rows:
            current_id = condition_id
        current_rows.append(row)
        previous_frame = frame
    if current_rows:
        groups.append((current_id, current_rows))
    return groups


def _condition_summary_rows(
    rows: Sequence[Mapping[str, object]],
) -> list[dict[str, object]]:
    """Build one Excel-ready summary row per simulated condition."""
    summaries: list[dict[str, object]] = []
    for _condition_id, condition_rows in _simulation_groups(rows):
        first = condition_rows[0]
        frame_count = len(condition_rows)
        squat_rows = [row for row in condition_rows if row.get("phase") == "isometrique"]
        if not squat_rows:
            squat_rows = [min(condition_rows, key=lambda row: _number(row, "com_y_m"))]

        support_values = [_number(row, "support_point_x_m") for row in condition_rows]
        outside_functional = sum(
            1
            for row in condition_rows
            if not bool(row.get("support_point_in_functional_base"))
        )
        outside_geometric = sum(
            1
            for row in condition_rows
            if not bool(row.get("support_point_in_geometric_base"))
        )
        over_limit = sum(
            1
            for row in condition_rows
            if any(
                bool(row.get(f"{joint}_utilization_exceeds_capacity"))
                for joint in JOINTS
            )
        )

        summary: dict[str, object] = {
            column: first.get(column) for column in SUMMARY_COLUMNS[:18]
        }
        summary.update(
            {
                "schema_version": SCHEMA_VERSION,
                "frames": frame_count,
                "squat_com_x_m": _mean(squat_rows, "com_x_m"),
                "squat_cop_x_m": _mean(squat_rows, "support_point_x_m"),
                "support_point_label": first.get("support_point_label"),
                "zmp_x_min_m": min(support_values),
                "zmp_x_max_m": max(support_values),
                "zmp_excursion_m": max(support_values) - min(support_values),
                "zmp_outside_support_frames": outside_functional,
                "zmp_outside_support_percent": (
                    100.0 * outside_functional / frame_count
                ),
                "cop_outside_foot_frames": outside_geometric,
                "cop_outside_foot_percent": (
                    100.0 * outside_geometric / frame_count
                ),
                "over_limit_frames": over_limit,
                "peak_grf_y_N": max(
                    abs(_number(row, "grf_y_N")) for row in condition_rows
                ),
            }
        )

        undefined_events: list[tuple[Mapping[str, object], str]] = []
        defined_events: list[tuple[float, Mapping[str, object], str]] = []
        for joint in JOINTS:
            utilizations = [
                float(row[f"{joint}_utilization_ratio"])
                for row in condition_rows
                if row.get(f"{joint}_utilization_ratio") is not None
            ]
            for row in condition_rows:
                ratio = row.get(f"{joint}_utilization_ratio")
                torque = _number(row, f"{joint}_torque_Nm")
                if ratio is None and abs(torque) > 0.0:
                    undefined_events.append((row, joint))
                elif ratio is not None:
                    defined_events.append((float(ratio), row, joint))
            summary.update(
                {
                    f"{joint}_peak_abs_torque_Nm": max(
                        abs(_number(row, f"{joint}_torque_Nm"))
                        for row in condition_rows
                    ),
                    f"{joint}_peak_abs_torque_body_mass_normalized_Nm_kg": max(
                        abs(
                            _number(
                                row,
                                f"{joint}_torque_body_mass_normalized_Nm_kg",
                            )
                        )
                        for row in condition_rows
                    ),
                    f"{joint}_peak_abs_power_W": max(
                        abs(_number(row, f"{joint}_power_W"))
                        for row in condition_rows
                    ),
                    f"{joint}_peak_utilization_ratio": (
                        max(utilizations) if utilizations else None
                    ),
                    f"{joint}_peak_utilization_percent": (
                        100.0 * max(utilizations) if utilizations else None
                    ),
                }
            )

        if undefined_events:
            limiting_row, limiting_joint = undefined_events[0]
            maximum_ratio: float | None = None
            exceeds_capacity = True
        elif defined_events:
            maximum_ratio, limiting_row, limiting_joint = max(
                defined_events, key=lambda item: item[0]
            )
            exceeds_capacity = maximum_ratio > 1.0
        else:
            maximum_ratio = None
            limiting_row = None
            limiting_joint = None
            exceeds_capacity = False
        summary.update(
            {
                "maximum_utilization_ratio": maximum_ratio,
                "maximum_utilization_percent": (
                    None if maximum_ratio is None else 100.0 * maximum_ratio
                ),
                "limiting_joint": limiting_joint,
                "limiting_frame": (
                    None if limiting_row is None else limiting_row.get("frame")
                ),
                "limiting_time_s": (
                    None if limiting_row is None else limiting_row.get("time_s")
                ),
                "limiting_phase": (
                    None if limiting_row is None else limiting_row.get("phase")
                ),
                "exceeds_capacity": exceeds_capacity,
                "undefined_capacity_events": len(undefined_events),
            }
        )
        summaries.append(summary)
    return summaries


def _project(
    rows: Sequence[Mapping[str, object]], columns: Sequence[str]
) -> list[list[object | None]]:
    return [[row.get(column) for column in columns] for row in rows]


_INVALID_WORKSHEET_CHARACTERS = re.compile(r'[\\/*?:\[\]<>|"\x00-\x1f]')
_WINDOWS_RESERVED_NAMES = re.compile(
    r"^(con|prn|aux|nul|com[1-9]|lpt[1-9])$", re.IGNORECASE
)


def _worksheet_name(
    condition_id: object,
    used_names: set[str],
    simulation_index: int,
) -> str:
    """Build a safe, unique Excel sheet name of at most 31 characters."""
    raw_name = "" if condition_id is None else str(condition_id)
    base = _INVALID_WORKSHEET_CHARACTERS.sub("_", raw_name)
    base = re.sub(r"\s+", " ", base).strip(" .'_")
    if not base:
        base = f"Simulation {simulation_index}"
    if _WINDOWS_RESERVED_NAMES.fullmatch(base):
        base = f"_{base}"
    base = base[:31].rstrip(" .'") or f"Simulation {simulation_index}"

    candidate = base
    suffix_index = 2
    while candidate.casefold() in used_names:
        suffix = f" ({suffix_index})"
        stem = base[: 31 - len(suffix)].rstrip(" .'")
        candidate = f"{stem}{suffix}"
        suffix_index += 1
    used_names.add(candidate.casefold())
    return candidate


def _ordered_row_columns(
    rows: Sequence[Mapping[str, object]],
) -> tuple[str, ...]:
    """Return the complete row schema in stable first-seen order."""
    return tuple(dict.fromkeys(column for row in rows for column in row))


def workbook_tables(
    rows: Sequence[Mapping[str, object]],
) -> OrderedDict[str, dict[str, object]]:
    """Build the student workbook: summary, combined frames and simulations."""
    if not rows:
        raise ValueError("L'export requiert au moins une ligne de résultats.")
    versioned = add_schema_version(rows)
    frame_columns = _ordered_row_columns(versioned)
    simulations = _simulation_groups(versioned)
    tables: OrderedDict[str, dict[str, object]] = OrderedDict()
    tables[SUMMARY_SHEET] = {
        "columns": list(SUMMARY_COLUMNS),
        "rows": _project(_condition_summary_rows(versioned), SUMMARY_COLUMNS),
    }
    tables[COMBINED_SHEET] = {
        "columns": list(frame_columns),
        "rows": _project(versioned, frame_columns),
    }

    used_names = {
        SUMMARY_SHEET.casefold(),
        COMBINED_SHEET.casefold(),
        DEFINITIONS_SHEET.casefold(),
    }
    for simulation_index, (condition_id, simulation_rows) in enumerate(
        simulations, start=1
    ):
        sheet_name = _worksheet_name(condition_id, used_names, simulation_index)
        tables[sheet_name] = {
            "columns": list(frame_columns),
            "rows": _project(simulation_rows, frame_columns),
        }

    definitions = []
    for csv_name, columns in (
        ("csv_standard", STANDARD_CSV_COLUMNS),
        ("csv_full", frame_columns),
    ):
        for column in columns:
            definition = column_definition(column)
            definitions.append(
                [
                    SCHEMA_VERSION,
                    csv_name,
                    column,
                    definition.unit,
                    definition.definition,
                    definition.sign_convention,
                    definition.status,
                ]
            )
    for table_name, columns in (
        (SUMMARY_SHEET, SUMMARY_COLUMNS),
        (COMBINED_SHEET, frame_columns),
        ("Simulation", frame_columns),
    ):
        for column in columns:
            definition = column_definition(column)
            definitions.append(
                [
                    SCHEMA_VERSION,
                    table_name,
                    column,
                    definition.unit,
                    definition.definition,
                    definition.sign_convention,
                    definition.status,
                ]
            )
    tables[DEFINITIONS_SHEET] = {
        "columns": [
            "schema_version",
            "table",
            "column",
            "unit",
            "definition",
            "sign_convention",
            "status",
        ],
        "rows": definitions,
    }
    return tables


def missing_dictionary_columns(tables: Mapping[str, Mapping[str, object]]) -> set[str]:
    """Return data columns that do not have an entry in the definitions table."""
    defined = {
        row[2] for row in tables[DEFINITIONS_SHEET]["rows"]  # type: ignore[index]
    }
    exported = {
        column
        for name, table in tables.items()
        if name != DEFINITIONS_SHEET
        for column in table["columns"]  # type: ignore[index]
    }
    return exported - defined


def _artifact_runtime() -> tuple[Path, Path]:
    """Locate Node and the artifact-tool modules without hard-coding a workstation."""
    node_candidates = [
        Path(value)
        for value in (os.environ.get("SQUAT_GUI_NODE"), shutil.which("node"))
        if value
    ]
    codex_dependencies = (
        Path.home() / ".cache/codex-runtimes/codex-primary-runtime/dependencies/node"
    )
    node_candidates.extend(
        (
            codex_dependencies / "bin/node",
            codex_dependencies / "bin/node.exe",
        )
    )
    module_candidates = [
        Path(value) for value in (os.environ.get("SQUAT_GUI_NODE_MODULES"),) if value
    ]
    module_candidates.extend(
        (
            Path.cwd() / "node_modules",
            codex_dependencies / "node_modules",
        )
    )
    node = next(
        (candidate for candidate in node_candidates if candidate.is_file()), None
    )
    modules = next(
        (
            candidate
            for candidate in module_candidates
            if (candidate / "@oai/artifact-tool").exists()
        ),
        None,
    )
    if node is None or modules is None:
        raise RuntimeError(
            "Export Excel indisponible: Node.js et @oai/artifact-tool sont requis. "
            "Définir SQUAT_GUI_NODE et SQUAT_GUI_NODE_MODULES si nécessaire."
        )
    return node, modules


def _link_artifact_modules(link: Path, modules: Path) -> None:
    """Expose the bundled Node modules in a temporary build directory."""
    try:
        link.symlink_to(modules, target_is_directory=True)
        return
    except OSError as error:
        if os.name != "nt":
            raise RuntimeError(
                f"Impossible de lier les modules Artifact Tool: {error}"
            ) from error
    completed = subprocess.run(
        ["cmd.exe", "/d", "/c", "mklink", "/J", str(link), str(modules)],
        check=False,
        capture_output=True,
        text=True,
    )
    if completed.returncode != 0:
        detail = completed.stderr.strip() or completed.stdout.strip()
        raise RuntimeError(
            f"Impossible de créer la jonction Artifact Tool: {detail}"
        )


def _write_xlsx_artifact(
    path: str | Path,
    rows: Sequence[Mapping[str, object]],
    *,
    preview_directory: str | Path | None = None,
) -> dict[str, object]:
    """Write the canonical tables with Artifact Tool when its runtime exists."""
    output = Path(path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    node, modules = _artifact_runtime()
    payload = {"schema_version": SCHEMA_VERSION, "tables": workbook_tables(rows)}
    builder = Path(__file__).with_name("build_workbook.mjs")
    if not builder.exists():
        raise RuntimeError(f"Constructeur de classeur introuvable: {builder}")

    with tempfile.TemporaryDirectory(prefix="squat-gui-xlsx-") as temporary:
        work = Path(temporary)
        local_builder = work / builder.name
        shutil.copy2(builder, local_builder)
        _link_artifact_modules(work / "node_modules", modules)
        payload_path = work / "payload.json"
        payload_path.write_text(
            json.dumps(payload, ensure_ascii=False, allow_nan=False),
            encoding="utf-8",
        )
        report_path = work / "report.json"
        command = [
            str(node),
            str(local_builder),
            str(payload_path),
            str(output),
            str(report_path),
        ]
        if preview_directory is not None:
            command.append(str(Path(preview_directory).expanduser().resolve()))
        completed = subprocess.run(
            command,
            cwd=work,
            check=False,
            capture_output=True,
            text=True,
        )
        # Artifact Tool 2.8 can terminate Node abnormally during Windows
        # teardown after preview rendering, despite having completed every
        # awaited write.  The per-run report is written last and therefore
        # provides a stronger completion signal than that teardown code.
        if completed.returncode != 0 and not (
            report_path.exists() and output.exists()
        ):
            detail = completed.stderr.strip() or completed.stdout.strip()
            raise RuntimeError(f"Échec de l'export Excel: {detail}")
        if not report_path.exists():
            raise RuntimeError("Échec de l'export Excel: rapport de validation absent.")
        report = json.loads(report_path.read_text(encoding="utf-8"))
        report["writer"] = "artifact-tool"
        return report


def _excel_number_format(column: str) -> str | None:
    if re.search(r"(^|_)(time|delta_time|duration).*_s$", column):
        return "0.000"
    if re.search(r"(_m|_m_s|_m_s2|_kg_m|_kg_m2|_N|_Nm|_W)$", column):
        return "0.000000"
    if re.search(r"(_deg|_deg_s|_deg_s2|_percent)$", column):
        return "0.000"
    return None


def _write_xlsx_openpyxl(
    path: str | Path,
    rows: Sequence[Mapping[str, object]],
) -> dict[str, object]:
    """Write the canonical workbook without an external Node.js runtime."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError(
            "Export Excel indisponible: installer openpyxl>=3.1 ou fournir "
            "Node.js avec @oai/artifact-tool."
        ) from error

    output = Path(path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    tables = workbook_tables(rows)
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="245B4A")
    alternate_fill = PatternFill("solid", fgColor="EAF2EE")
    header_font = Font(color="FFFFFF", bold=True)
    subtle_border = Side(style="thin", color="D8E1DD")
    long_text_columns = {
        "anthropometry_scaling_rule",
        "scaling_rule",
        "contact_source",
        "support_point_source",
        "capacity_model",
        "capacity_source",
        "definition",
        "sign_convention",
    }

    for sheet_index, (name, table) in enumerate(tables.items(), start=1):
        sheet = workbook.create_sheet(name)
        sheet.sheet_view.showGridLines = False
        columns = list(table["columns"])
        table_rows = list(table["rows"])
        sheet.append(columns)
        for values in table_rows:
            sheet.append(list(values))

        sheet.row_dimensions[1].height = 34
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )

        for row_index in range(2, sheet.max_row + 1):
            for cell in sheet[row_index]:
                border = copy(cell.border)
                border.bottom = subtle_border
                cell.border = border
                if row_index % 2 == 0:
                    cell.fill = alternate_fill

        for column_index, column in enumerate(columns, start=1):
            letter = get_column_letter(column_index)
            number_format = _excel_number_format(column)
            is_long_text = column in long_text_columns
            maximum_width = 48 if name == DEFINITIONS_SHEET or is_long_text else 24
            measured_width = max(
                len(str(sheet.cell(row=row_index, column=column_index).value or ""))
                for row_index in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[letter].width = min(
                max(measured_width + 2, 11), maximum_width
            )
            for row_index in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                if number_format and isinstance(cell.value, (int, float)):
                    cell.number_format = number_format
                if is_long_text:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        sheet.freeze_panes = "C2"
        if sheet.max_row >= 2:
            reference = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            excel_table = Table(
                displayName=f"SquatTable{sheet_index}", ref=reference
            )
            excel_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(excel_table)

    workbook.save(output)
    check = load_workbook(output, read_only=False, data_only=False)
    try:
        formula_errors = [
            cell.value
            for sheet in check.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
            and cell.value in {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
        ]
        sheets = check.sheetnames
    finally:
        check.close()
    return {
        "sheets": sheets,
        "formulaErrors": formula_errors,
        "writer": "openpyxl",
    }


def write_xlsx(
    path: str | Path,
    rows: Sequence[Mapping[str, object]],
    *,
    preview_directory: str | Path | None = None,
) -> dict[str, object]:
    """Write the canonical workbook, with an autonomous openpyxl fallback."""
    writer = os.environ.get("SQUAT_GUI_XLSX_WRITER", "auto").strip().lower()
    if writer not in {"auto", "artifact-tool", "openpyxl"}:
        raise ValueError(
            "SQUAT_GUI_XLSX_WRITER doit valoir auto, artifact-tool ou openpyxl."
        )
    if writer in {"auto", "artifact-tool"}:
        try:
            return _write_xlsx_artifact(
                path, rows, preview_directory=preview_directory
            )
        except RuntimeError:
            if writer == "artifact-tool":
                raise
    return _write_xlsx_openpyxl(path, rows)
