"""Physical Excel writers for the canonical Squat GUI workbook model."""

from __future__ import annotations

from copy import copy
import json
import os
from pathlib import Path
import shutil
import subprocess
import tempfile
from typing import Mapping

from .workbook_model import (
    DEFINITIONS_SHEET,
    LONG_TEXT_COLUMNS,
    WorkbookTable,
    excel_number_format,
)


def _artifact_runtime() -> tuple[Path, Path]:
    """Locate Node and the artifact-tool modules without hard-coding a workstation."""
    node_candidates = [
        Path(value)
        for value in (os.environ.get("SQUAT_GUI_NODE"), shutil.which("node"))
        if value
    ]
    codex_dependencies = (
        Path.home() / ".cache/codex-runtimes/codex-primary-runtime/dependencies/node"
    )
    node_candidates.extend(
        (
            codex_dependencies / "bin/node",
            codex_dependencies / "bin/node.exe",
        )
    )
    module_candidates = [
        Path(value) for value in (os.environ.get("SQUAT_GUI_NODE_MODULES"),) if value
    ]
    module_candidates.extend(
        (
            Path.cwd() / "node_modules",
            codex_dependencies / "node_modules",
        )
    )
    node = next(
        (candidate for candidate in node_candidates if candidate.is_file()), None
    )
    modules = next(
        (
            candidate
            for candidate in module_candidates
            if (candidate / "@oai/artifact-tool").exists()
        ),
        None,
    )
    if node is None or modules is None:
        raise RuntimeError(
            "Export Excel indisponible: Node.js et @oai/artifact-tool sont requis. "
            "Définir SQUAT_GUI_NODE et SQUAT_GUI_NODE_MODULES si nécessaire."
        )
    return node, modules


def _link_artifact_modules(link: Path, modules: Path) -> None:
    """Expose the bundled Node modules in a temporary build directory."""
    try:
        link.symlink_to(modules, target_is_directory=True)
        return
    except OSError as error:
        if os.name != "nt":
            raise RuntimeError(
                f"Impossible de lier les modules Artifact Tool: {error}"
            ) from error
    completed = subprocess.run(
        ["cmd.exe", "/d", "/c", "mklink", "/J", str(link), str(modules)],
        check=False,
        capture_output=True,
        text=True,
    )
    if completed.returncode != 0:
        detail = completed.stderr.strip() or completed.stdout.strip()
        raise RuntimeError(
            f"Impossible de créer la jonction Artifact Tool: {detail}"
        )


def write_xlsx_artifact(
    path: str | Path,
    tables: Mapping[str, WorkbookTable],
    *,
    schema_version: str,
    builder: str | Path,
    preview_directory: str | Path | None = None,
) -> dict[str, object]:
    """Write already-normalized workbook tables with Artifact Tool."""
    output = Path(path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    node, modules = _artifact_runtime()
    payload = {"schema_version": schema_version, "tables": tables}
    builder_path = Path(builder)
    if not builder_path.exists():
        raise RuntimeError(f"Constructeur de classeur introuvable: {builder_path}")

    with tempfile.TemporaryDirectory(prefix="squat-gui-xlsx-") as temporary:
        work = Path(temporary)
        local_builder = work / builder_path.name
        shutil.copy2(builder_path, local_builder)
        _link_artifact_modules(work / "node_modules", modules)
        payload_path = work / "payload.json"
        payload_path.write_text(
            json.dumps(payload, ensure_ascii=False, allow_nan=False),
            encoding="utf-8",
        )
        report_path = work / "report.json"
        command = [
            str(node),
            str(local_builder),
            str(payload_path),
            str(output),
            str(report_path),
        ]
        if preview_directory is not None:
            command.append(str(Path(preview_directory).expanduser().resolve()))
        completed = subprocess.run(
            command,
            cwd=work,
            check=False,
            capture_output=True,
            text=True,
        )
        # Artifact Tool 2.8 can terminate Node abnormally during Windows
        # teardown after preview rendering, despite having completed every
        # awaited write. The report is written last and is the stronger signal.
        if completed.returncode != 0 and not (
            report_path.exists() and output.exists()
        ):
            detail = completed.stderr.strip() or completed.stdout.strip()
            raise RuntimeError(f"Échec de l'export Excel: {detail}")
        if not report_path.exists():
            raise RuntimeError("Échec de l'export Excel: rapport de validation absent.")
        report = json.loads(report_path.read_text(encoding="utf-8"))
        report["writer"] = "artifact-tool"
        return report


def write_xlsx_openpyxl(
    path: str | Path, tables: Mapping[str, WorkbookTable]
) -> dict[str, object]:
    """Write already-normalized workbook tables with openpyxl."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError(
            "Export Excel indisponible: installer openpyxl>=3.1 ou fournir "
            "Node.js avec @oai/artifact-tool."
        ) from error

    output = Path(path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="245B4A")
    alternate_fill = PatternFill("solid", fgColor="EAF2EE")
    header_font = Font(color="FFFFFF", bold=True)
    subtle_border = Side(style="thin", color="D8E1DD")

    for sheet_index, (name, table) in enumerate(tables.items(), start=1):
        sheet = workbook.create_sheet(name)
        sheet.sheet_view.showGridLines = False
        columns = list(table["columns"])
        table_rows = list(table["rows"])
        sheet.append(columns)
        for values in table_rows:
            sheet.append(list(values))

        sheet.row_dimensions[1].height = 34
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )

        for row_index in range(2, sheet.max_row + 1):
            for cell in sheet[row_index]:
                border = copy(cell.border)
                border.bottom = subtle_border
                cell.border = border
                if row_index % 2 == 0:
                    cell.fill = alternate_fill

        for column_index, column in enumerate(columns, start=1):
            letter = get_column_letter(column_index)
            number_format = excel_number_format(column)
            is_long_text = column in LONG_TEXT_COLUMNS
            maximum_width = 48 if name == DEFINITIONS_SHEET or is_long_text else 24
            measured_width = max(
                len(str(sheet.cell(row=row_index, column=column_index).value or ""))
                for row_index in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[letter].width = min(
                max(measured_width + 2, 11), maximum_width
            )
            for row_index in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                if number_format and isinstance(cell.value, (int, float)):
                    cell.number_format = number_format
                if is_long_text:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        sheet.freeze_panes = "C2"
        if sheet.max_row >= 2:
            reference = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            excel_table = Table(
                displayName=f"SquatTable{sheet_index}", ref=reference
            )
            excel_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(excel_table)

    workbook.save(output)
    check = load_workbook(output, read_only=False, data_only=False)
    try:
        formula_errors = [
            cell.value
            for sheet in check.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
            and cell.value in {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
        ]
        sheets = check.sheetnames
    finally:
        check.close()
    return {
        "sheets": sheets,
        "formulaErrors": formula_errors,
        "writer": "openpyxl",
    }
