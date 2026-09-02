import math
import os
import tempfile
import unittest
from unittest.mock import patch
from pathlib import Path
from zipfile import ZipFile

from squat_gui.cli import Condition, simulate_condition
from squat_gui.export_schema import (
    COMBINED_SHEET,
    DEFINITIONS_SHEET,
    SCHEMA_VERSION,
    STANDARD_CSV_COLUMNS,
    SUMMARY_COLUMNS,
    SUMMARY_SHEET,
    csv_export_rows,
    missing_dictionary_columns,
    workbook_tables,
    write_xlsx,
)


class ExportSchemaTests(unittest.TestCase):
    @staticmethod
    def rows() -> list[dict[str, object]]:
        condition = Condition(
            condition_id="contrat",
            load_percent_bw=20.0,
            subject_profile="homme",
            bar_position="front",
            wedge_20_deg=True,
            shank_percent=2.0,
            thigh_percent=-1.0,
            trunk_percent=3.0,
            anthropometry_mode="morphotype recalibre",
            duration_excentrique_s=2.0,
            duration_isometrique_s=0.5,
            duration_concentrique_s=2.0,
            q_segment_deg=(22.0, -58.0, 20.0),
            torque_preset="Anderson actif x2",
            max_torques={"cheville": 222.0, "genou": 380.0, "hanche": 376.0},
            angle_adapt=True,
            velocity_adapt=True,
            frames=3,
            backend="analytical",
        )
        rows, _summary = simulate_condition(condition)
        return rows

    def test_every_excel_column_has_definition_and_units(self) -> None:
        rows = self.rows()
        tables = workbook_tables(rows)
        self.assertEqual(
            list(tables),
            [SUMMARY_SHEET, COMBINED_SHEET, "contrat", DEFINITIONS_SHEET],
        )
        self.assertEqual(missing_dictionary_columns(tables), set())
        self.assertTrue(
            all(
                row[0] == SCHEMA_VERSION
                for row in tables[COMBINED_SHEET]["rows"]
            )
        )
        csv_standard_definitions = {
            row[2]: row
            for row in tables[DEFINITIONS_SHEET]["rows"]
            if row[1] == "csv_standard"
        }
        csv_full_definitions = {
            row[2]: row
            for row in tables[DEFINITIONS_SHEET]["rows"]
            if row[1] == "csv_full"
        }
        self.assertEqual(tuple(csv_standard_definitions), STANDARD_CSV_COLUMNS)
        self.assertEqual(set(csv_full_definitions), set(rows[0]))
        self.assertEqual(
            csv_full_definitions["zmp_in_support"][6], "compatibilité legacy"
        )
        self.assertEqual(
            csv_full_definitions["cheville_contact_Nm"][6],
            "compatibilité legacy",
        )
        self.assertEqual(
            csv_full_definitions["cheville_mass_acceleration_Nm"][6],
            "canonique",
        )
        self.assertEqual(
            tables[COMBINED_SHEET]["columns"], tables["contrat"]["columns"]
        )
        self.assertEqual(tables[COMBINED_SHEET]["rows"], tables["contrat"]["rows"])

    def test_simulation_sheet_names_are_safe_unique_and_keep_duplicate_ids(self) -> None:
        base_rows = self.rows()
        condition_ids = [
            "Synthèse",
            "essai/chargé",
            "essai?chargé",
            "x" * 40,
            "x" * 39 + "y",
            "",
            "CON",
            "dup",
            "dup",
        ]
        rows = [
            {**row, "condition_id": condition_id}
            for condition_id in condition_ids
            for row in base_rows
        ]

        tables = workbook_tables(rows)
        simulation_names = list(tables)[2:-1]
        self.assertEqual(
            simulation_names,
            [
                "Synthèse (2)",
                "essai_chargé",
                "essai_chargé (2)",
                "x" * 31,
                "x" * 27 + " (2)",
                "Simulation 6",
                "_CON",
                "dup",
                "dup (2)",
            ],
        )
        self.assertTrue(all(len(name) <= 31 for name in simulation_names))
        self.assertEqual(len(tables[SUMMARY_SHEET]["rows"]), len(condition_ids))
        self.assertEqual(
            len(tables[COMBINED_SHEET]["rows"]), len(base_rows) * len(condition_ids)
        )
        duplicate_summary_ids = [
            row[1]
            for row in tables[SUMMARY_SHEET]["rows"]
            if row[1] == "dup"
        ]
        self.assertEqual(duplicate_summary_ids, ["dup", "dup"])

    def test_standard_csv_contract_is_explicit_and_full_mode_is_compatible(self) -> None:
        rows = self.rows()
        standard = csv_export_rows(rows)
        complete = csv_export_rows(rows, mode="full")

        self.assertEqual(tuple(standard[0]), STANDARD_CSV_COLUMNS)
        self.assertEqual(standard[0]["schema_version"], SCHEMA_VERSION)
        self.assertEqual(standard[0]["frames"], 3)
        self.assertIn("cheville_torque_body_mass_normalized_Nm_kg", standard[0])
        self.assertIn("support_point_in_functional_base", standard[0])
        self.assertNotIn("cheville_mass_acceleration_Nm", standard[0])
        self.assertNotIn("foot_weighted_com_x_kg_m", standard[0])
        self.assertIn("cheville_mass_acceleration_Nm", complete[0])
        self.assertIn("foot_weighted_com_x_kg_m", complete[0])

    def test_excel_summary_contains_requested_student_metrics(self) -> None:
        rows = self.rows()
        table = workbook_tables(rows)[SUMMARY_SHEET]
        self.assertEqual(tuple(table["columns"]), SUMMARY_COLUMNS)
        self.assertEqual(len(table["rows"]), 1)
        summary = dict(zip(table["columns"], table["rows"][0]))

        support_values = [float(row["support_point_x_m"]) for row in rows]
        self.assertEqual(summary["frames"], len(rows))
        self.assertAlmostEqual(
            float(summary["zmp_excursion_m"]),
            max(support_values) - min(support_values),
        )
        self.assertAlmostEqual(
            float(summary["cheville_peak_abs_torque_Nm"]),
            max(abs(float(row["cheville_torque_Nm"])) for row in rows),
        )
        self.assertIn(summary["limiting_joint"], ("cheville", "genou", "hanche"))

    def test_global_com_is_the_sum_of_exported_segment_contributions(self) -> None:
        segments = ("foot", "shank", "thigh", "trunk", "bar")
        for row in self.rows():
            total_mass = float(row["total_mass_kg"])
            weighted_x = sum(
                float(row[f"{segment}_weighted_com_x_kg_m"]) for segment in segments
            )
            weighted_y = sum(
                float(row[f"{segment}_weighted_com_y_kg_m"]) for segment in segments
            )
            self.assertTrue(
                math.isclose(
                    weighted_x / total_mass, float(row["com_x_m"]), abs_tol=1e-12
                )
            )
            self.assertTrue(
                math.isclose(
                    weighted_y / total_mass, float(row["com_y_m"]), abs_tol=1e-12
                )
            )

    def test_artifact_xlsx_contains_student_workflow_and_no_formula_error(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "squat.xlsx"
            previews = Path(temporary) / "previews"
            with patch.dict(
                os.environ, {"SQUAT_GUI_XLSX_WRITER": "artifact-tool"}, clear=False
            ):
                try:
                    report = write_xlsx(
                        output, self.rows(), preview_directory=previews
                    )
                except RuntimeError as error:
                    self.skipTest(f"Runtime Artifact Tool indisponible: {error}")
            self.assertEqual(report["writer"], "artifact-tool")
            self.assertEqual(report["formulaErrors"], [])
            self.assertEqual(
                report["sheets"],
                [SUMMARY_SHEET, COMBINED_SHEET, "contrat", DEFINITIONS_SHEET],
            )
            self.assertTrue(
                all((previews / f"{name}.png").exists() for name in report["sheets"])
            )
            with ZipFile(output) as archive:
                workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
            for name in report["sheets"]:
                self.assertIn(f'name="{name}"', workbook_xml)

    def test_openpyxl_fallback_is_autonomous_and_preserves_contract(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "fallback.xlsx"
            with patch.dict(
                os.environ, {"SQUAT_GUI_XLSX_WRITER": "openpyxl"}, clear=False
            ):
                report = write_xlsx(output, self.rows())
            self.assertEqual(report["writer"], "openpyxl")
            self.assertEqual(report["formulaErrors"], [])
            self.assertEqual(
                report["sheets"],
                [SUMMARY_SHEET, COMBINED_SHEET, "contrat", DEFINITIONS_SHEET],
            )

            from openpyxl import load_workbook

            workbook = load_workbook(output, read_only=False, data_only=False)
            try:
                self.assertEqual(workbook.sheetnames, report["sheets"])
                self.assertEqual(workbook.sheetnames[0], SUMMARY_SHEET)
                self.assertEqual(workbook[SUMMARY_SHEET].freeze_panes, "C2")
                self.assertEqual(workbook[COMBINED_SHEET].freeze_panes, "C2")
                self.assertEqual(workbook["contrat"].freeze_panes, "C2")
                self.assertEqual(
                    workbook[COMBINED_SHEET]["E2"].value,
                    self.rows()[0]["delta_time_s"],
                )
                self.assertEqual(
                    workbook[COMBINED_SHEET]["E2"].number_format, "0.000"
                )
                self.assertTrue(workbook[COMBINED_SHEET].tables)
                self.assertTrue(workbook["contrat"].tables)
                self.assertEqual(
                    workbook[SUMMARY_SHEET]["A1"].fill.fgColor.rgb[-6:], "245B4A"
                )
            finally:
                workbook.close()

    def test_auto_falls_back_when_artifact_runtime_is_absent(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "auto.xlsx"
            with (
                patch.dict(
                    os.environ, {"SQUAT_GUI_XLSX_WRITER": "auto"}, clear=False
                ),
                patch(
                    "squat_gui.export_schema._write_xlsx_artifact",
                    side_effect=RuntimeError("Node absent"),
                ),
            ):
                report = write_xlsx(output, self.rows())
            self.assertEqual(report["writer"], "openpyxl")
            self.assertTrue(output.exists())


if __name__ == "__main__":
    unittest.main()
